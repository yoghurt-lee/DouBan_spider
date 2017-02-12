#coding:utf-8
'''
爬取豆瓣的书的信息
'''

import os
import random
from time import sleep
import urllib

from bs4 import BeautifulSoup
from openpyxl import Workbook

from Downloader import downloader



path = 'data/'
def book_spider(book_tag):
    page_num = 0
    book_list = []
    retry_times = 0
    title_set = set({}) #防止书籍重复
    while(True):
        url = 'https://www.douban.com/tag/'+urllib.quote(book_tag)+'/book?start='+str(page_num*15)
        sleep(random.uniform(2,6)) #随机睡眠时间
        html = downloader(url)
        if html is None: continue
        text = str(html)
        soup = BeautifulSoup(text) #通过BeautifulSoup来做匹配
        list_soup = soup.find('div',{'class':'mod book-list'})
        retry_times+=1
        if list_soup==None and retry_times<100:
            continue
        elif list_soup==None or len(list_soup)<=1: 
            break
        for book_info in list_soup.findAll("dd"):
            title = book_info.find('a',{'class':'title'}).string.strip()
            if(title in title_set): continue
            title_set.add(title)
            desc = book_info.find('div',{'class':'desc'}).string.strip()
            #book_url = book_info.find('a',{'class':'title'}).get('href')
            list_desc = desc.split('/')
            try:
                author_info = '/'.join(list_desc[:-3])
            except:
                author_info = '不详'
            try:
                public_info = '/'.join(list_desc[-3:-1])
            except:
                public_info = '不详'
            try:
                price_info = list_desc[-1]
            except:
                price_info = '不详'
            try:
                rating_info = book_info.find('span',{'class':'rating_nums'}).string.strip()
            except:
                rating_info = '0.0'
            book_list.append([title,author_info,public_info,price_info,rating_info])
            retry_times=0 #reset 0
        page_num+=1
        print "Now crawling %d pages..." % page_num
    return book_list
        
def start_spider(book_tags_list):
    book_lists = []
    for book_tag in book_tags_list:
        book_list = book_spider(book_tag) #获得某个标签的书的内容
        book_list = sorted(book_list,key=lambda x: x[4],reverse=True) #按照书的评分的降序排列
        book_lists.append(book_list)
    return book_lists

def print_into_excel(book_tag_lists, book_lists):
    wb=Workbook()
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))
    for i in range(len(book_tag_lists)):
        ws[i].append(['序号','书名','评分  ','作者/译者','定价','出版信息'])
        count = 0
        for book_list in book_lists[i]:
            ws[i].append([count,book_list[0],book_list[4],book_list[1],book_list[3],book_list[2]])
            count+=1
    save_path = path+'tag'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    if os.path.exists(save_path):
        print u"该标签文件已存在..."
        return
    wb.save(save_path)
    
    
    
    